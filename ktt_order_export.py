import os
import sys
import pandas as pd
pd.set_option('display.unicode.east_asian_width', True) #设置输出右对齐
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter



class KttList:
    def __init__(self):
        ktt_config=os.path.join(os.path.dirname(__file__),'web','config','ktt.config')
        with open(ktt_config,'r',encoding='utf-8') as f:
            self.ktt_config=json.loads(f.read())


        # col_names_config=os.path.join(os.path.dirname(__file__),'web','config','ktt.config')
        with open(self.ktt_config['col_config_fn'],'r',encoding='utf-8') as f:
            self.config=json.loads(f.read())
        # print(self.config.keys())

    def dataframe_output(self,supplier,info):
        df=pd.DataFrame(data=info['buy_recs'],columns=self.config[supplier]['col_names'])   
        # df.append(info,ignore_index=True)
        col_map=self.config[supplier]['col_map']
        for pos,itm in enumerate(self.config[supplier]['col_names']):
            if col_map[itm]=='spec':
                spec_col=pos
        
        return {'order_df':df,'fn_str':info['spec']+' '+str(df.shape[0])+'件','count':df.shape[0],'spec_col':spec_col}

    def read_web(self,supplier,sender_name,sender_tel,spec,lst):
        col_names=self.config[supplier]['col_names']
        col_map=self.config[supplier]['col_map']
        _name_lines=lst.strip().split('\n')
        name_lines=list(filter(None,_name_lines))
        name_lines=[itm.strip() for itm in name_lines]

        recs=[]        
        for itm in name_lines:
            #中英文逗号均可处理
            try:
                recs.append([itm.split('，')[0],int(itm.split('，')[1])])
            except:
                recs.append([itm.split(',')[0],int(itm.split(',')[1])])
        
        #按config文件中的排序 
        buy_recs=[]
        for itm in recs:
            itm_detail=itm[0].split(' ')
            for i in range(itm[1]):
                buy_rec=[]
                for key,value in col_map.items():

                    if value=='rcv_name':
                        buy_rec.append(itm_detail[0])
                    elif value=='rcv_tel':
                        buy_rec.append(itm_detail[1])
                    elif value=='rcv_adr':
                        buy_rec.append(itm_detail[2])
                    elif value=='spec':
                        buy_rec.append(spec)
                    elif value=='sender_name':
                        buy_rec.append(sender_name)
                    elif value=='sender_tel':
                        buy_rec.append(sender_tel)
                    else:
                        buy_rec.append(value)

                buy_recs.append(buy_rec)
        
        return {'buy_recs':buy_recs,'spec':spec}
            
    
    def output(self,supplier,sender_name,sender_tel,spec,lst):
        info=self.read_web(supplier=supplier,sender_name=sender_name,sender_tel=sender_tel,spec=spec,lst=lst)
        res=self.dataframe_output(supplier=supplier,info=info)

        #self.dataframe_output返回格式： {'order_df':df,'fn_str':info['spec']+' '+str(df.shape[0])+'件','count':df.shape[0],'spec_col':spec_col}
        return res

    def multi_spec_output(self,supplier,sender_name,sender_tel,odrs,save='yes',save_cfg=['团团好果','20240109','01','脆蜜金桔'],save_dir='e:\\temp\\ktt\\exp_order'):
        dfs=[]
        fn_str=''
        desc_str=''
        total_count=0
        for odr in odrs:
            res=self.output(supplier=supplier,sender_name=sender_name,sender_tel=sender_tel,spec=odr[0],lst=odr[1])
            df=res['order_df']
            fn_str=fn_str+res['fn_str']+' '
            desc_str=desc_str+res['fn_str']+'\r\n'
            dfs.append(df)
            total_count=total_count+res['count']            
        
        out_df=pd.concat(dfs)

        # print(out_df)
        # print(fn_str.strip())

        if save=='yes':
            fn='-'.join(save_cfg)+'-'+fn_str.strip()+'.xlsx'
            save_fn=os.path.join(save_dir,fn)
            out_df.to_excel(save_fn,sheet_name='导入模板',index=False)

            if len(odrs)>1:

                xlsx_desc='共{}件\n其中：\n{}'.format(str(total_count),desc_str)
                
            else:
                xlsx_desc=desc_str

            #将文件名记录到临时文件中
            tmp_fn=os.path.join('e:\\temp\\ktt\\exp_order','newfn.tmp')
            with open (tmp_fn, 'w') as f:
                f.write(fn)

            self.xlsx_format(save_fn,desc=xlsx_desc,desc_row=total_count+4,desc_col=res['spec_col']+1)
        
        return out_df

    def xlsx_format(self,xlsx,desc,desc_row,desc_col):
        wb=openpyxl.load_workbook(xlsx)
        ws=wb.active

        # for cell in ws[1]:
            
        #     font=Font(size=13,bold=True)
        #     cell.font=font

        font=Font(size=13,bold=True)
        ws['A1'].font=font

        ws.cell(desc_row,desc_col).value=desc

        for cls in ['A','B','C','D','F','G','H','I']:
            cell=ws[cls+'1']
            font=Font(color='FF0000',bold=True)
            cell.font=font

        
        #调整列宽
        for cell in ws[1]:
            max_length = 0
            column_letter = get_column_letter(cell.column)
            # print(column_letter,cell.value)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

            if column_letter=='C':
                adjusted_width = (max_length + 15) * 3.2
            else:
                adjusted_width = (max_length + 5) * 3.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(xlsx)
        print('修改格式完成')


            
        



if __name__=='__main__':
    p=KttList()
    # p.list_line(good_type='脆蜜金桔',收件人姓名='王中',收件人手机='13800001234',收件人地址='广西南宁市',商品名称='脆蜜果王家庭装【二】三斤',单号='',快递='',发件人姓名='团团好果',发件人手机='13717710616')
    # p.one_line('脆蜜金桔',['王中','13800001234','广西南宁市','脆蜜果王家庭装【二】三斤','','','团团好果','13717710616'])
    # infos=[['王中','13800001234','广西南宁市','脆蜜果王家庭装【二】三斤','','','团团好果','13717710616']]
    # infos=[['王中','13800001234','广西南宁市','脆蜜果王家庭装【二】三斤','','','团团好果','13717710616'],['王二','13800001234','广西南宁市','脆蜜果王家庭装【二】三斤','','','团团好果','13717710616'],['王三','13800001234','广西南宁市','脆蜜果王家庭装【二】三斤','','','团团好果','13717710616']]
    # p.list_lines(good_type='脆蜜金桔',infos=infos) 
    # res=p.dataframe_output('脆蜜金桔',infos)
    # print(res)
    # p.one_line(good_type='脆蜜金桔',rcv_name='王中',rcv_tel='13800001234',rcv_adr='广西南宁市',good_name='脆蜜果王家庭装【二】三斤',sender_name='团团好果',sender_tel='13717710616')
    # sender_name='团团好果'
    # sender_tel='13717710616'
    # spec0='脆蜜果王家庭装【二】三斤'
    # lst0='''
    # 王中 13707710060 广西南宁市，2
    # 王一 13899238593 广西玉林市，3
    # 余女士 15921806296 上海市上海市宝山区纬地路88弄15号102,4
    # '''
    # spec1='滑皮家庭装【二】五斤'
    # lst1='''
    # 李中 13707710060 广西南宁市，1
    # 李一 13899238593 广西玉林市，2
    # 狗 15921806296 上海市上海市宝山区纬地路88弄15号102,2
    # '''

    # odrs=[
    #     [spec0,lst0]
    # ]

    # # res=p.output(good_type='脆蜜金桔',sender_name=sender_name,sender_tel=sender_tel,spec=spec,lst=lst)
    # # print(res)
    # p.multi_spec_output(supplier='姐妹',sender_name=sender_name,sender_tel=sender_tel,odrs=odrs,save='yes',save_cfg=['团团好果','20240109','01','脆蜜金桔'],save_dir='e:\\temp')
