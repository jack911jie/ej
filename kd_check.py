import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__),'modules'))
import format_transfer
import read_config
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
# 创建Chrome选项
chrome_options = Options()
chrome_options.add_argument('--headless')  # 设置为无界面模式
import re
import time
import pandas as pd
pd.set_option('display.unicode.east_asian_width', True) #设置输出右对齐
import copy
import win32com.client
import openpyxl
from openpyxl.styles import Font, Color,PatternFill,Alignment
from openpyxl.utils import get_column_letter
import numpy as np
import json


class FruitKd:
    def __init__(self,chromedriver_path):
        if chromedriver_path:
            selenium_service = Service(chromedriver_path)
            self.driver=webdriver.Chrome(service=selenium_service)
        else:
            pass

        #无界面模式
        # self.driver=webdriver.Chrome(chromedriver_path,options=chrome_options)

    def close_chrome_driver(self):
        self.driver.quit()

    #第一次获取物流记录
    def connect_kd_web(self,phone_number,url,phn_digits=4):
        # 打开网页
        self.driver.get(url)
        time.sleep(1)

        # 设置要输入的号码
        phone_number_4 = str(phone_number)[phn_digits*-1:] # 替换成你想要的号码

        # 找到收件人输入框并输入号码
        try:
            recipient_input = self.driver.find_element(By.ID,"query_str")
            recipient_input.send_keys(phone_number_4)

            # 提交表单
            self.driver.find_element(By.ID,"submit_product_query").click()

            # 获取结果
            result = self.driver.page_source

            # print(result.text)

            # 关闭浏览器
            # self.driver.quit()
        except Exception as err:
            print('连接查询网址时出错')
            result=''
        # finally:
        #     result=''
        #     self.driver.quit()

        return result

    #如有重复记录的获取方法2
    def connect_kd_web_2(self,phone_number,url,phn_digits):
        # 打开网页
        self.driver.get(url)
        time.sleep(1)

        # 设置要输入的号码
        phone_number_4 = str(phone_number)[phn_digits*-1:] # 替换成你想要的号码

        # 找到收件人输入框并输入号码
        try:
            # recipient_input = self.driver.find_element(By.ID,"query_str")
            # recipient_input.send_keys(phone_number_4)

            # 提交表单
            # self.driver.find_element(By.ID,"submit_product_query").click()

            # 获取结果
            result = self.driver.page_source

            # print(result.text)

            # 关闭浏览器
            # self.driver.quit()
        except Exception as err:
            print('连接查询网址时出错')
            result=''
        # finally:
        #     result=''
        #     self.driver.quit()

        return result

    def guoyuan_xlsx_style(self,xls):
            wb=openpyxl.load_workbook(xls)
            ws=wb.active

            # for cell in ws[1]:
                
            #     font=Font(size=13,bold=True)
            #     cell.font=font

            font=Font(size=13,bold=True)
            ws['A1'].font=font

            for cls in ['A','B','C','D','F','G','H','I']:
                cell=ws[cls+'1']
                font=Font(color='FF0000',bold=True)
                cell.font=font

            
            #调整列宽
            for cell in ws[1]:
                max_length = 0
                column_letter = get_column_letter(cell.column)
                # print(column_letter,cell.value)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

                if column_letter=='J':
                    adjusted_width = (max_length + 15) * 2.2
                else:
                    adjusted_width = (max_length + 5) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(xls)
            print('修改格式完成')

    def guige_to_filename(self,dict):
        df_gp=dict.groupby(['规格'])
        df_gp_counts=df_gp.count()
        fn_dict=pd.DataFrame(df_gp_counts['数量']).reset_index().set_index('规格')['数量'].to_dict()
        fn_out=''
        #将规格中的阿拉伯数字转为中文
        for key,value in fn_dict.items():
            fn_out+=format_transfer.number_to_chinese(int(key[:-2]))+key[-2:]+str(value)+'件 '
        fn_out=fn_out.strip()
        return fn_out

    def many_or_single_result(self,res,phn,phn_digits,check_date,keyword,ptn):
        soup=BeautifulSoup(res,'html.parser')

        if '查询到多条记录' in res:
            print('{} 有多条记录，正在匹配日期及姓名。'.format(phn))
            # if '顺丰' in kd_name:
            check_dates=[str(x) for x in check_date]
            date_txts=[x[4:6]+'月'+x[6:]+'日' for x in check_dates]

            li_elements = soup.select('li[role="presentation"]')
            # <li role="presentation" class="active"><a href="http://kd.dh.cx/df66d/7329/0"> 06月16日（单号：776260474339358） </a></li>
            urls=[]
            for li in li_elements:
                a_tag=li.find('a')
                if a_tag:
                    for date_txt in date_txts:
                        if date_txt in a_tag.text:
                            url=a_tag['href']                    
                            urls.append(url)
            
            # print(urls)

            ress=[]
            for url in urls:
                # url=url.split('//')[-1]
                res=self.connect_kd_web_2(phone_number=phn,url=url,phn_digits=phn_digits)
                tmp_res=self.deal_result(res=res,keyword=keyword,ptn=ptn)
                ress.append(tmp_res[0])

            result=ress
        else:
            print('{} 有一条记录'.format(phn))
            result=self.deal_result(res,keyword=keyword,ptn=ptn)
                
        
        return result
           
    def deal_result(self,res,keyword='物流单号',ptn='\d{15}'):
        # 处理结果
        # 这里需要根据具体情况来解析和提取你想要的信息

         # 输出结果

        soup=BeautifulSoup(res,'html.parser')
        li_elements = soup.find_all('li')

        

        # 遍历每个li元素
        info=[]
        for li in li_elements:
            # print(li.get_text())
            tt=''
            # 检查li元素的文本内容是否包含目标字符
            try:
                
                if keyword in li.get_text():
                    # 找到了包含目标字符的li元素
                    kd_code_txt=li
                    code=re.findall(ptn,str(li.text))[0]
                    # code=re.findall(ptn,str(li.text))
                    tt=tt+code
            
    
                if '收件人：' in li.get_text():
                    kd_receiver_txt=li

                    name=str(li.text).split('：')[-1]
                    name=name.strip()

                    tt=tt+name
                    # break

                if len(tt)>0:
                    info.append(tt)
            except Exception as err_code:
                print(err_code)
                pass
        
        info_0=info[0::2]
        info_1=info[1::2]

        
        # result=[n_ph[info_0[x],info_1[x]] for x in range(len(info_0))]

        result=[]
        for x in range(len(info_0)):
            n_ph={}
            n_ph[info_0[x]]=info_1[x] #顺丰用            
            # n_ph[info_1[x]]=info_0[x] #申通用
            result.append(n_ph)        

        return result

    def batch_phone_number(self,phn_name_list=['15678892330阿晓','17853297329李君'],url='http://kd.dh.cx/df66d',check_date=[20230703],page_keyword='单号',phn_digits=4,ptn=r'\d{15}'):
        kd_id_list=[]
        for phn_name in phn_name_list:
            phn=phn_name[:11]
            name=phn_name[11:]
            id_get=self.connect_kd_web(phone_number=phn,url=url,phn_digits=phn_digits)

            # print('id_get:',id_get)
            # with open('d:\\py\\test\\res.html', 'r', encoding='utf-8') as fhtml:
            #     id_get = fhtml.read()
            if id_get:
                kddh=self.many_or_single_result(res=id_get,phn=phn,phn_digits=phn_digits,check_date=check_date,keyword=page_keyword,ptn=ptn)

                # print(kddh)
                if kddh:
                    for kd in kddh:
                        # print(kd)
                        try:
                            kd_id_list.append([str(phn),name,str(kd[name])])
                        except Exception as e:
                            # print('{} 号码下未找到 {} 的订单号。'.format(phn,name))
                            kd_id_list.append([str(phn),name,np.nan])
                else:
                    #  print('{} 号码下未找到 {} 的订单号。'.format(phn,name))
                     kd_id_list.append([str(phn),name,np.nan])

        return kd_id_list

    def read_order_excel(self,xls='e:\\temp\\ejj\\团购群\\订单\\给果园的订单\\团团好果06.17订单-01.xlsx'):
        df_read=pd.read_excel(xls)
        df_new=df_read[['收件人手机','收件人姓名']]
        df=copy.deepcopy(df_new)
        df['联系电话收货人']=df_new['收件人手机'].astype(str)+df_new['收件人姓名']
        
        return df['联系电话收货人'].tolist()

    def read_dl_order_excel(self,xls='e:\\temp\\ejj\\团购群\\订单\\wuliu2023-06-17 21_27_45.xlsx.xlsx'):
        df_dl_order=pd.read_excel(xls,sheet_name='订单统计')
        df_dl_order_2=df_dl_order[['联系电话','收货人']]
        df_dl_order_new=df_dl_order_2.copy()
        df_dl_order_new['联系电话收货人']=df_dl_order_2['联系电话'].astype(str)+df_dl_order_2['收货人']

        return df_dl_order_new['联系电话收货人'].tolist()

    def order_to_guoyuan(self,dl_xls='E:\\temp\\ejj\\团购群\\订单\\0614导出.xlsx',output_dir='e:\\temp\\ejj\\团购群\\给果园的订单',
                        check_ice_bag='yes',ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config',
                        expand_accounts='yes',exp='yes'):

        #检查文件名：
        if not re.match(r'\d{8}-.*\d{0,3}(斤|两|个)装-导出订单-\d\d.xlsx',dl_xls.split('\\')[-1]):
            exit('文件名格式错误。正确格式类似：20230708-台农10斤装-导出订单-01.xlsx')

        df_order=pd.read_excel(dl_xls,sheet_name='订单列表')
        df_order_out=df_order.copy()
        df_order_out=df_order_out[['收货人','联系电话','详细地址','商品','订单号','订单金额']]
        df_order_out.rename(columns={'收货人':'收件人姓名','联系电话':'收件人手机','详细地址':'收件人地址','商品':'品名'},inplace=True)
        df_order_out['收件人手机']=df_order_out['收件人手机'].apply(lambda x: str(x))
        df_order_out['规格']=df_order_out['品名'].apply(lambda x: re.findall(r'(\d\d斤装|\d斤装)',x)[0])
        df_order_out['数量']=df_order_out['品名'].apply(lambda x: int(str(x).split('+')[-1]))
        df_order_out['发件人姓名']='团团好果'
        df_order_out['发件人手机']='18077796420'
        df_order_out['发件人电话']=''
        df_order_out['发件人地址']=''
        df_order_out['发件人单位']=''
        df_order_out['收件人电话']=''
        df_order_out['收件人单位']=''

        # df_order_out['规格']=''
        
        
        
        df_order_out['备注']=''
        df_order_out['代收金额']=''
        df_order_out['到付金额']=''
        df_order_out['规格和地址']=df_order_out['规格']+'%'+df_order_out['收件人地址']
        df_order_out['重量小计']=df_order_out['规格'].apply(lambda x: int(x[:-2]))*df_order_out['数量']

        #计算总重量:
        total_wt=df_order_out['重量小计'].sum()
        # print(total_wt)
    
        # print(df_order_out['规格和地址'].str.split('%'))

        df_order_out=df_order_out[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','备注','订单号','代收金额','到付金额','规格和地址','订单金额']]
        if check_ice_bag=='yes':
            # df_order_out['冰袋数量']=df_order_out['规格和地址'].apply(lambda x: self.ice_bag_number(addr=x.split('%')[1],spec=x.str.split('%')[0],ice_bag_fn=ice_bag_fn))
            df_order_out['冰袋数量']=df_order_out['规格和地址'].apply(lambda x: self.ice_bag_number_df(addr_spec=x,ice_bag_fn=ice_bag_fn))
        else:
            df_order_out['冰袋数量']=0
        
            #row['收件人地址'].apply(lambda x),spec=row['规格'],ice_bag_fn=ice_bag_fn
      
        if expand_accounts=='yes':
            df_repeated=df_order_out.loc[df_order_out.index.repeat(df_order_out['数量'])]
            df_repeated=df_repeated.reset_index(drop=True)
            df_repeated['数量']=1

            df_repeated['重复订单']=df_repeated.duplicated(subset='订单号',keep='first')
            df_repeated.loc[df_repeated['重复订单'],'订单金额']=0
            df_repeated.drop('重复订单',axis=1,inplace=True)

            df_res=df_repeated
        else:
            df_order_out['数量']=df_order_out['品名'].apply(lambda x: int(str(x).split('+')[-1]))
            df_res=df_order_out

        fn_out=self.guige_to_filename(df_res)

        

        # print(df_repeated)
        # exp：显示内容并保存，'yes'参数用于处理单独的一张表，没有多表合并的情况
        if exp=='yes':
            xlsname=dl_xls.split('\\')[-1].split('.')[0].split('-')
            datetxt,num=xlsname[0],xlsname[2]
            #正则匹配出水果名字，如 台农10斤，结果为 台农
            frt_name=re.findall(r'(.*?)(?=\d)',dl_xls.split('\\')[-1].split('-')[1])[0]

            fn='团团好果'+datetxt[4:6]+'.'+datetxt[6:]+'订单'+'-'+frt_name+'-'+fn_out+'.xlsx'
            dir_fn=os.path.join(output_dir,fn)
            df_res=df_res.drop(columns=['规格和地址']) #去掉“规格和地址”再保存
            df_res=df_res[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','冰袋数量','备注','订单号','代收金额','到付金额','订单金额']]
            df_res.to_excel(dir_fn,index=False)
            print('{} 导出完成'.format(fn))
            self.guoyuan_xlsx_style(xls=dir_fn)
            os.startfile(output_dir)
        else:
            df_res=df_res.drop(columns=['规格和地址'])#去掉“规格和地址”再输出
            # df_res=df_res[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','备注','订单号','代收金额','到付金额']]
    
        try:
            result_res=df_res[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','冰袋数量','备注','订单号','代收金额','到付金额','订单金额']]
        except Exception as err:
            result_res=pd.DataFrame()
            print('导出dataframe错误',err)
            # pass
            # result_res=df_res[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','备注','订单号','代收金额','到付金额','订单金额']]
        
        return {'orders':result_res,'total_wt':total_wt}

    def multi_order_to_guoyuan(self,date=20230618,fn_keyword_ptn='台农\d{0,4}斤',input_dir='E:\\temp\\ejj\\团购群\\订单',output_dir='e:\\temp\\ejj\\团购群\\订单\\给果园的订单',
                                check_ice_bag='yes',ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config',
                                expand_accounts='yes',save_each_exp='no'):
        date=str(date)
        if fn_keyword_ptn:
            frt_name=fn_keyword_ptn.split('\\')[0]
        else:
            fn_keyword_ptn='水果'
            frt_name=fn_keyword_ptn
        try:
            fns=[]
            for fn in os.listdir(input_dir):
                if re.match(str(date)+'-'+fn_keyword_ptn+'-导出订单-\d\d.xlsx',fn):
                    fns.append(os.path.join(input_dir,fn))

            dfs=[]
            total_wt=0
            for fn in fns:
                df_res=self.order_to_guoyuan(dl_xls=fn,output_dir=output_dir,check_ice_bag=check_ice_bag,ice_bag_fn=ice_bag_fn,
                                            expand_accounts=expand_accounts,exp=save_each_exp)

                if df_res['orders'].shape[0]>0:
                    dfs.append(df_res['orders'])
                    total_wt+=df_res['total_wt']

            df_concat=pd.concat(dfs)

            fn_out=self.guige_to_filename(df_concat)


            fn='团团好果'+date[4:6]+'.'+date[6:]+'-'+'订单（'+str(len(fns))+'个订单合并）-'+frt_name+'-'+fn_out+'.xlsx'
            wt_txt='（'+str(len(fns))+'个订单一共{}斤）'.format(str(total_wt))

            #0614导出.xlsx
            dir_fn=os.path.join(output_dir,fn)
            df_concat.to_excel(dir_fn,index=False)
            print('\n{} 导出完成\n{}'.format(fn,wt_txt))
            self.guoyuan_xlsx_style(xls=dir_fn)
            os.startfile(output_dir)

            return df_concat

        except Exception as e:
            print('错误：',e)

    def write_xlsx_back_kd(self,input_xls='e:\\temp\\ejj\\团购群\\订单\\给果园的订单\\团团好果06.17订单-01.xlsx',
                            out_dir='e:\\temp\\ejj\\团购群\\订单\\带物流信息的回传文件',
                            url='http://kd.dh.cx/df66d',
                            kd_name='申通快递',
                            check_date=20230703,
                            method='download',
                            page_keyword='单号',phn_digits=4,ptn=r'\d{15}'):
        if method=='download':
            #根据method值获取输入表内容
            df_input=pd.read_excel(input_xls)
            title=df_input.columns.tolist()

            #有的时候下载的表格中，列名 物流单号（必填）的括号格式每次不一样，故从下载的表格中提取这个列名，以保证改名或合并时不出错。
            wuliudh_txt=title[2]
            #生成客户手机及姓名列表
            exp_list=self.read_dl_order_excel(xls=input_xls)
            #获取客户物流单号
            res=self.batch_phone_number(phn_name_list=exp_list,url=url,check_date=check_date,page_keyword=page_keyword,phn_digits=phn_digits,ptn=ptn)
            
            df_write=pd.DataFrame(data=res,columns=['联系电话','收货人',wuliudh_txt])

            #匹配结果去重
            # print(df_write)
            df_write.drop_duplicates(subset=['联系电话','收货人',wuliudh_txt],inplace=True)

            print('\n\n------------------------\n以下为匹配结果：\n',df_write)
            df_kd=df_write.dropna(how='any',subset=[wuliudh_txt])

            

            #如无快递单号的df
            if df_kd.shape[0]==0:
                return '未返回有效快递单号'
                # pass                
            else:
                df_input['物流公司（必填）']=kd_name
                # kd_id_dic=df_kd.set_index('收件人手机')['物流单号'].to_dict()
                # df_input['物流单号（必填）']=df_input['联系电话'].apply(lambda x:kd_id_dic.get(str(x),''))
                # df_kd.rename(columns={'收件人手机':'联系电话','收件人姓名':'收货人','物流单号':'物流单号（必填）'},inplace=True)

                df_input['联系电话']=df_input['联系电话'].apply(lambda x: str(x))
                df_tmp=df_input.copy()
                df_tmp=pd.merge(df_kd,df_input,on =['收货人','联系电话'],how='outer')
                # print(df_input)
                df_tmp.drop(wuliudh_txt+'_y',axis=1,inplace=True)
                df_tmp.rename(columns={wuliudh_txt+'_x':wuliudh_txt},inplace=True)
                df_tmp.drop_duplicates(subset=['收货人','联系电话','物流单号（必填）'],inplace=True)

                if not os.path.exists(out_dir):
                    os.makedirs(out_dir)
                out_fn=os.path.join(out_dir,input_xls.split('\\')[-1][:-5]+'-已写入物流单号待上传.xlsx')
                # print(df_tmp)
                df_tmp=df_tmp[title]
                df_tmp.to_excel(out_fn,index=False)

                os.startfile(out_dir)
                return '\n写入待回传文件完成'

        else:
            pass

        self.close_chrome_driver()

    def ice_bag_number(self,addr,spec,ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config'):
        ice_bag=2
        if ice_bag_fn:
            provinces=read_config.read_json(fn=ice_bag_fn)
            for pro in provinces['one_icebag']:
                # if pro in addr:
                if re.match('^'+pro+'.*',addr) and spec=='10斤装':
                    ice_bag=1
                    break

        return ice_bag

    def ice_bag_number_df(self,addr_spec,ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config'):
        ice_bag=2
        spec,addr=addr_spec.split('%')

        if ice_bag_fn:
            provinces=read_config.read_json(fn=ice_bag_fn)
            for pro in provinces['one_icebag']:
                # if pro in addr:
                if re.match('^'+pro+'.*',addr):
                    ice_bag=1
                    break

        if spec=='5斤装':
            ice_bag=1

        return ice_bag

    #师院百香果格式
    def shiyuan_format_to_wuliu(self,wuliu_company='中通快递',
                                    guoyuan_back_fn='E:\\temp\\ejj\\团购群\\订单\\果园返单\\订单查询_20230824181755_8lbs4oek4800.xls',
                                    ktt_dl_wuliu_empty_fn='E:\\temp\\ejj\\团购群\\订单\\wuliu2023-08-24 21_01_40.xlsx.xlsx',
                                    output_to_upload_dir='E:\\temp\\ejj\\团购群\\订单\\带物流信息的回传文件'):
        df_back=pd.read_excel(guoyuan_back_fn)
        df_dl=pd.read_excel(ktt_dl_wuliu_empty_fn)
        #发现有的时候快团团下载的文件，有关物流公司和物流单号的括号写法不一致，通过以下匹配获取。
        for col_name in df_dl.columns:
            if re.match(r'^物流公司.*',col_name):
                wuliu_comp_name=col_name
            if re.match(r'^物流单号.*',col_name):
                wuliu_code=col_name
        
        df_dl[wuliu_comp_name]=wuliu_company
        df_dl[wuliu_code]=df_dl['联系电话'].apply(lambda x: str(df_back[df_back['收件人手机号']==x]['运单编号'].tolist()[0]))
        
        _to_upload_fn=ktt_dl_wuliu_empty_fn.split('\\')[-1].split('.')[0]+'已写入物流单号待上传.xlsx'
        to_upload_fn=os.path.join(output_to_upload_dir,_to_upload_fn)
        df_dl.to_excel(to_upload_fn,index=False)

        print('已写入快递单号')

class LocalProduct(FruitKd):
    def __init__(self,chromedriver_path):
        super().__init__(chromedriver_path)
        with open(os.path.join(os.path.dirname(__file__),'configs','ktt.config'),'r',encoding='utf-8') as f:
            self.sender_info=json.load(f)
        self.ice_bag_fn=r'd:\py\ej\configs\ktt_ice_bag.config'

        with open(self.sender_info['ktt_col_config'],'r',encoding='utf-8') as f_colnames:
            self.col_names=json.load(f_colnames)

        with open(self.sender_info['ktt_col_map_config'],'r',encoding='utf-8') as f_map:
            self.col_map_config=json.load(f_map)
            

        
    def read_order_excel(self,fn):
        #去除快团团的密码，原来设置为7788.去除后才能读取
        xlsApp=win32com.client.DispatchEx('Excel.Application')
        xlsApp.EnableEvents=False
        xlsApp.DisplayAlerts=False
        xlwb=xlsApp.Workbooks.Open(fn,UpdateLinks=False, ReadOnly=False, Format=None, Password=7788, WriteResPassword='')
        xlwb.SaveAs(fn,None,"","")
        xlsApp.Quit()
        
        df=pd.read_excel(fn,sheet_name='商品列表',engine='openpyxl')
        df['联系电话']=df['联系电话'].astype(str)
        df=df[['订单号','商品','规格','数量','收货人','联系电话','详细地址','订单金额']]
        return df
    
    def deal_order(self,good_name,fn,format='fruit',check_ice_bag='no',expand_accounts='yes'):
        df=self.read_order_excel(fn)
        if format=='local_product':
            df['发货人']=self.sender_info['发货人姓名']
        elif format=='fruit':
            df['发货人']=self.sender_info['发货人昵称']
        else:
            df['发货人']=self.sender_info['发货人姓名']

        df['发货人身份证号']=self.sender_info['发货人身份证号']
        df['发货人电话']=self.sender_info['发货人电话']
        if good_name:
            df['商品']=good_name

        try:
            ptn=r'\d+g'
            df['规格']=df['规格'].apply(lambda x: re.findall(ptn,x)[0])
        except:
            pass

        if expand_accounts=='yes':
            df_repeated=df.loc[df.index.repeat(df['数量'])]
            df_repeated=df_repeated.reset_index(drop=True)
            df_repeated['数量']=1

            df_repeated['重复订单']=df_repeated.duplicated(subset='收货人',keep='first')
            df_repeated.loc[df_repeated['重复订单'],'订单金额']=0
            df_repeated.drop('重复订单',axis=1,inplace=True)

            df_res=df_repeated
        else:
            df_res=df

        
        if format=='local_product':
            df_res=df_res[['商品','规格','数量','收货人','联系电话','详细地址','发货人','发货人身份证号','发货人电话','订单金额']]
        elif format=='fruit':
            df_res.rename(columns={'发货人':'发件人姓名','发货人电话':'发件人手机','收货人':'收件人姓名','联系电话':'收件人手机','详细地址':'收件人地址','商品':'品名',},inplace=True)
            # df_res.rename(columns={'发货人电话':'发件人手机'},inplace=True)
            #发件人姓名	发件人手机	发件人电话	发件人地址	发件人单位	收件人姓名	收件人手机	收件人电话	收件人地址	收件人单位	品名	规格	数量	冰袋数量	备注	订单号	代收金额	到付金额
            df_res['发件人电话']=''
            df_res['收件人电话']=''
            df_res['发件人地址']=''
            df_res['发件人单位']=''
            df_res['收件人单位']=''
            
            df_res['备注']=''
            df_res['代收金额']=''
            df_res['到付金额']=''
            df_res['规格和地址']=df_res['规格']+'%'+df_res['收件人地址']

            if check_ice_bag=='yes':
                # df_order_out['冰袋数量']=df_order_out['规格和地址'].apply(lambda x: self.ice_bag_number(addr=x.split('%')[1],spec=x.str.split('%')[0],ice_bag_fn=ice_bag_fn))
                df_res['冰袋数量']=df_res['规格和地址'].apply(lambda x: self.ice_bag_number_df(addr_spec=x,ice_bag_fn=self.ice_bag_fn))
            else:
                df_res['冰袋数量']=0
            

            df_res=df_res[['发件人姓名','发件人手机','发件人电话','发件人地址','发件人单位','收件人姓名','收件人手机','收件人电话','收件人地址','收件人单位','品名','规格','数量','冰袋数量','备注','订单号','代收金额','到付金额','订单金额']]
        else:
            df_res=df_res[['商品','规格','数量','收货人','联系电话','详细地址','发货人','发货人身份证号','发货人电话','订单金额']]
        
        
        return df_res

    def send_to_producer(self,out_dir,out_fn_prefix,supplier,batch,good_name,fn,good_format='local_product',check_ice_bag='no',expand_accounts='yes',open_dir='yes'):
        df=self.deal_order(good_name=good_name,fn=fn,format=good_format,check_ice_bag=check_ice_bag,expand_accounts=expand_accounts)

         #修改规格，快团团规格转换为提供给发货商的规格
        try:
            df['规格']=df['规格'].apply(lambda x: self.alter_spec(supplier,x))
        except:
            pass

        #如有默认设置，则填入。
        try:
            for col in self.col_names[supplier]['col_default']:
                df[col]=self.col_names[supplier]['col_default'][col]
        except:
            pass

        #按规格统计
        df_grp=df.groupby('规格')['数量'].sum()

        #修改规格，快团团规格转换为提供给发货商的规格
        speciality_str = ' '.join([f'{self.alter_spec(supplier,spec)} {count}件' for spec, count in df_grp.items()])
        desc='\n'.join([f'{self.alter_spec(supplier,spec)} {count}件' for spec, count in df_grp.items()])
        total_sum=str(df['数量'].sum())

        #根据规格的种类数量写描述
        values_count=len(df['规格'].value_counts())
        if values_count>1:
            desc= f'共{total_sum}件\n其中：'+desc


        #修改格式时，将描述写入商品列下4行，定位商品单元格
        good_name_row=df.shape[0]+4
        good_name_col=int(self.col_names[supplier]['col_goodname'])+1


        if not df.empty:
            date_input=fn.split('\\')[-1].split('-')[0]
            out_fn=os.path.join(out_dir,f'{out_fn_prefix}-{date_input}-{batch}-{good_name}-{speciality_str}.xlsx')

            supplier_df=self.alter_supplier_format(supplier=supplier,df=df)
            # print(supplier_df)

            supplier_df.to_excel(out_fn,sheet_name='团团好果发货单',index=False)

            #修改格式
            self.xlsx_format(out_fn,desc=desc,desc_row=good_name_row,desc_col=good_name_col)

            if open_dir=='yes':
                os.startfile(out_dir)
            print(f'完成。文件名：{out_fn}')
            return {'res':'ok','data':df,'filename':out_fn}
        else:
            print('数据为空')
            return {'res':'failed','error':'empty data input'}


    def xlsx_format(self,xlsx,desc,desc_row,desc_col):
        wb=openpyxl.load_workbook(xlsx)
        ws=wb.active

        # for cell in ws[1]:
            
        #     font=Font(size=13,bold=True)
        #     cell.font=font

        font=Font(size=13,bold=True)
        ws['A1'].font=font

        ws.cell(desc_row,desc_col).value=desc
        font=Font(bold=True)
        ws.cell(desc_row,desc_col).font=font
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.cell(desc_row,desc_col).fill=yellow_fill
        ws.cell(desc_row,desc_col).alignment = Alignment(wrapText=True)

        for cls in ['A','B','C','D','F','G','H','I']:
            cell=ws[cls+'1']
            font=Font(color='FF0000',bold=True)
            cell.font=font

        
        #调整列宽
        for cell in ws[1]:
            max_length = 0
            column_letter = get_column_letter(cell.column)
            # print(column_letter,cell.value)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

            if column_letter=='C':
                adjusted_width = (max_length + 15) * 3.2
            else:
                adjusted_width = (max_length + 5) * 3.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(xlsx)
        # print('修改格式完成')

    def alter_supplier_format(self,supplier,df):
        supplier_col_names=self.col_names[supplier]['col_names']

        #倒置map表的key和value
        # reversed_map={}
        # for key,value in self.col_names[supplier]['col_map2'].items():
        #     reversed_map[value]=key


        # col_names_map=self.col_names[supplier]['col_map2']
        # old_cols=list(df.columns)
        # print(old_cols)
        newdf=pd.DataFrame(columns=supplier_col_names)
        for col in supplier_col_names:
            # print(col, self.col_names[supplier]['col_map2'][col])
            try:
                newdf[col]=df[self.col_names[supplier]['col_map2'][col]]
            except:
                pass

        # print(newdf)
        return newdf


    def alter_spec(self,supplier,spec):
        try:
            new_spec=self.col_map_config[supplier][spec]
        except:
            new_spec=spec
        # print(new_spec)
        return new_spec


if __name__=='__main__':
    #龙眼干及其他水果发货
    p=LocalProduct(chromedriver_path='')
    res=p.send_to_producer(out_dir='E:\\temp\\ejj\\团购群\\订单\\给果园的订单',
                        out_fn_prefix='团团好果',
                        supplier='尚朴',
                        batch='02',
                        good_name='尚朴滑皮金桔',   
                        good_format='fruit',      
                        check_ice_bag='no',               
                        expand_accounts='yes',
                        fn='E:\\temp\\ejj\\团购群\\订单\\20240108-尚朴滑皮金桔-导出订单-01.xlsx')
    
    #参数说明：
    # out_dir：输出文件夹
    # out_fn_prefix：文件名的前缀
    # supplier： 供应商名称，如脆蜜的供应商有“姐妹”，“尚朴”。不同供应商的不同货物有不同的发货模板，对应的模板放在本地某个名为ktt_col.config的文件中
    # batch：批次。同一天有不同的批次，体现在导出的文件名中 。
    # good_name：商品名，出现在表格中的“商品名”列
    # good_format：商品形式，local_product-土特产,fruit-水果。根据不同的类型生成不同的表格，土特产，为邮政要求的格式，即有发货人的真实姓名、身份证号，水果，则发货人对应为config文件中的发货人昵称，通常为团团好果。
    # check_ice_bag：是否检查冰袋，通常只有荔枝需要设置为yes
    # expand_accounts：对于购买多于1件的商品，根据实际数量生成相应的记录。例如同一客户购买了3件，同一条记录生成3条，防止商家发货漏单。
    # fn：从快团团出导出并下载的表格
    # ！！！注意！！！快团团导出的订单可能有密码导致读取不出商品列表，需要手动在excel表中清除密码。


    
    ##测试版块
    #顺丰：快递单号 r'SF\d{13}'
    #申通：物流单号 r'\d{15}'
    # p=FruitKd(chromedriver_path='D:/Program Files (x86)/ChromeWebDriver/chromedriver')
    # res=p.batch_phone_number(phn_name_list=['13811776353冯勤'],url='http://kd.dh.cx/36cd9',check_date=[20230703,20230628],page_keyword='快递单号',phn_digits=11,ptn=r'SF\d{13}')
    # # res=p.batch_phone_number(phn_name_list=['17853297329李君'],url='http://kd.dh.cx/df66d',page_keyword='物流单号',phn_digits=4,ptn=r'\d{15}')
    # print(res)

 
    # 一、从快团团批量导入订单处理后生成给果园的订单。只能处理一个文件。
    # p=FruitKd(chromedriver_path='')
    # rs=p.order_to_guoyuan(dl_xls='E:\\temp\\ejj\\团购群\\订单\\20230619-台农10斤-导出订单-02.xlsx',
    #                             output_dir='e:\\temp\\ejj\\团购群\\订单\\给果园的订单',
    #                             check_ice_bag='yes',ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config',
    #                             expand_accounts='yes',
    #                             exp='yes')
    #参数说明：
    # dl_xls：从快团团批量导出的订单，文件名为：20230619-台农10斤装-导出订单-02.xlsx 的格式，校验格式为：\d{8}-.*\d{0,3}(斤|两|个)装-导出订单-\d\d.xlsx
    # output_dir：生成给果园的订单文件后存放的文件夹
    # expand_accounts：对于购买多于1件的商品，根据实际数量生成相应的记录。例如同一客户购买了3件，同一条记录生成3条，防止商家发货漏单。
    # exp：是否显示信息。（在多单导入时建议no，否则会显示很多信息）
    # print(rs)

    #批量处理同一天的不同订单，可将多个订单生成一个合并发货清单文件给果园。
    # p=FruitKd(chromedriver_path='')
    # rs=p.multi_order_to_guoyuan(date=20230630,
    #                             fn_keyword_ptn='台农\d{0,4}斤',
    #                             input_dir='E:\\temp\\ejj\\团购群\\订单',
    #                             output_dir='e:\\temp\\ejj\\团购群\\订单\\给果园的订单',
    #                             check_ice_bag='yes',ice_bag_fn='d:\\py\\ej\\configs\\ktt_ice_bag.config',
    #                             expand_accounts='yes',
    #                             save_each_exp='no')

    #参数说明：
    # dl_xls：从快团团批量导出的订单，文件名修改为：20230618-导出订单-02.xlsx 的格式
    # fn_keyword_ptn：文件夹中要匹配的文件名的正则表达，如：'台农\d{0,2}斤',
    # output_dir：生成给果园的订单文件后存放的文件夹
    # expand_accounts：对于购买多于1件的商品，根据实际数量生成相应的记录。例如同一客户购买了3件，同一条记录生成3条，防止商家发货漏单。
    # each_exp：处理到每个订单时，是否分别保存。默认为no，即只保存最后生成的合并订单。
    # print(rs)

    # 二、果园返单后，通过下载快团团模板文件查询快递单号并写入待上传文件
    # p=FruitKd(chromedriver_path='D:/Program Files (x86)/ChromeWebDriver/chromedriver')
    # res=p.write_xlsx_back_kd(input_xls='e:\\temp\\ejj\\团购群\\订单\\20230707-桂七物流-.xlsx',
    #                         out_dir='e:\\temp\\ejj\\团购群\\订单\\带物流信息的回传文件',
    #                         url='http://kd.dh.cx/bb505',
    #                         check_date=[20230708],
    #                         kd_name='中通快递',
    #                         method='download',
    #                         page_keyword='运单编号',phn_digits=4,ptn=r'\d{14}')
    # print(res)

    #参数说明：
    # input_xls：从快团团导出的待回传清单
    # output_dir：生成给果园的订单文件后存放的文件夹
    # url：查询地址，可能会经常改变
    # check_date:查询有多条记录时，只查询输入的check_date日期下的快递记录，check_date为8位数字日期格式的list
    # kd_name: 快递公司名称。写入生成的回传清单中。如：申通快递，顺丰快递
    # method: download——从快团团导出的待回传清单文件查询物流单号并生成回传文件以上传。目前仅支持这一选项。
    # page_keyword：不同的快递查询页面关键字不一样，目前已知：申通——物流单号，顺丰——快递单号
    # phn_digits：不同的查询页面要求的电话号码位数不一样，目前已知：申通——4位，顺丰——11位
    # ptn：匹配的快递单号模式，目前已知：申通——r'\d{15}'，顺丰——r'SF\d{13}'

    # 其他：申通查询http://kd.dh.cx/df66d，用手机号后4位，顺丰查询http://kd.dh.cx/36cd9，用整个手机号11位

    # 三、师院百香果返回的物流匹配快团团下载的待回传物流单号文件
    # p=FruitKd(chromedriver_path='')
    # p.shiyuan_format_to_wuliu(wuliu_company='中通快递',
    #                                 guoyuan_back_fn='E:\\temp\\ejj\\团购群\\订单\\果园返单\\订单查询_20230824181755_8lbs4oek4800.xls',
    #                                 ktt_dl_wuliu_empty_fn='E:\\temp\\ejj\\团购群\\订单\\wuliu2023-08-24 21_01_40.xlsx.xlsx',
    #                                 output_to_upload_dir='E:\\temp\\ejj\\团购群\\订单\\带物流信息的回传文件')




    